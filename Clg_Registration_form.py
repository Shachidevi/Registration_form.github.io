# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *
from tkinter import filedialog
import os
# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('C:\\Users\\Smart\\Desktop\\pythonproject\\sheetBook.xlsx')

# create the sheet object
sheet = wb.active


def excel():
	
	# resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 30
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 30
	sheet.column_dimensions['G'].width = 20
	sheet.column_dimensions['H'].width = 20
	sheet.column_dimensions['I'].width = 30
	

	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "Name"
	sheet.cell(row=1, column=2).value = "Contact Number"
	sheet.cell(row=1, column=3).value = "education"
	sheet.cell(row=1, column=4).value = "Institute"
	sheet.cell(row=1, column=5).value = "Email id"
	sheet.cell(row=1, column=6).value = "Address"
	sheet.cell(row=1, column=7).value = "Gender"
	sheet.cell(row=1, column=8).value = "Password"
	sheet.cell(row=1, column=9).value = "Re-Enter Password"


# Function to set focus (cursor)
def focus1(event):
	# set focus on the course_field box
	ContactNum_field.focus_set()


# Function to set focus
def focus2(event):
	# set focus on the sem_field box
	edu_field.focus_set()


# Function to set focus
def focus3(event):
	# set focus on the form_no_field box
	Institute_field.focus_set()


# Function to set focus
def focus4(event):
	# set focus on the contact_no_field box
        email_id_field.focus_set()


# Function to set focus
def focus5(event):
	# set focus on the email_id_field box
	address_field.focus_set()


# Function to set focus
def focus6(event):
	# set focus on the address_field box
	Gender_field.focus_set()

def focus7(event):
	# set focus on the address_field box
	password_field.focus_set()

def focus8(event):
	# set focus on the address_field box
	Re_enter_password_field.focus_set()




# Function for clearing the
# contents of text entry boxes
def clear():
	
	# clear the content of text entry box
	name_field.delete(0, END)
	ContactNum_field.delete(0, END)
	edu_field.delete(0, END)
	Institute_field.delete(0, END)
	email_id_field.delete(0, END)
	address_field.delete(0, END)
	Gender_field.delete(0, END)
	password_field.delete(0, END)
	Re_enter_password_field.delete(0, END)


# Function to take data from GUI 
# window and write to an excel file
def insert():
	
	# if user not fill any entry
	# then print "empty input"
	if (name_field.get() == "" and
		ContactNum_field.get() == "" and
		edu_field.get() == "" and
		Institute_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == "" and
		Gender_field.get() == "" and
                password_field.get() == "" and
                Re_enter_password_field.get() == ""):
			
		print("empty input")

	else:

		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
		current_row = sheet.max_row
		current_column = sheet.max_column

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		sheet.cell(row=current_row + 1, column=2).value = ContactNum_field.get()
		sheet.cell(row=current_row + 1, column=3).value = edu_field.get()
		sheet.cell(row=current_row + 1, column=4).value = Institute_field.get()
		sheet.cell(row=current_row + 1, column=5).value = email_id_field.get()
		sheet.cell(row=current_row + 1, column=6).value = address_field.get()
		sheet.cell(row=current_row + 1, column=7).value = Gender_field.get()
		sheet.cell(row=current_row + 1, column=8).value = password_field.get()
		sheet.cell(row=current_row + 1, column=9).value = Re_enter_password_field.get()

		# save the file
		wb.save('C:\\Users\\Smart\\Desktop\\pythonproject\\sheetBook.xlsx')

		# set focus on the name_field box
		name_field.focus_set()

		# call the clear() function
		clear()


    
# Driver code
if __name__ == "__main__":
	
	# create a GUI window
	root = Tk()

	# set the background colour of GUI window
	root.configure(background='gray')

	# set the title of GUI window
	root.title("registration form")

	# set the configuration of GUI window
	root.geometry("550x250")

	excel()

	# create a Form label
	heading = Label(root, text="Form", bg="light green")

	# create a Name label
	name = Label(root, text="Name: ", bg="gray")

	# create a Contact Number label
	ContactNum = Label(root, text="Contact Number: ", bg="gray")

	# create a Education label
	edu = Label(root, text="Education: ", bg="gray")

	# create a Gender. label
	Institute = Label(root, text="Institute: ", bg="gray")

	# create a Email-id. label
	email_id = Label(root, text="Email-id: ", bg="gray")

	# create a Address label
	address = Label(root, text="Address: ", bg="gray")

	# create a Country label
	Gender = Label(root, text="Gender: ", bg="gray")

	# browes the file from system
	password = Label(root, text = "Password: ", bg="gray")

	Re_enter_password = Label(root, text = "Re-enter Password: ", bg="gray")
     
        
        #button_exit = Button(window, 
                     #text = "Exit",
                     #command = exit) 

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	heading.grid(row=0, column=1)
	name.grid(row=1, column=0)
	ContactNum.grid(row=2, column=0)
	edu.grid(row=3, column=0)
	Institute.grid(row=4, column=0)
	email_id.grid(row=5, column=0)
	address.grid(row=6, column=0)
	Gender.grid(row=7, column=0)
	password.grid(row = 8, column=0)
	Re_enter_password.grid(row = 9, column=0)

	# create a text entry box
	# for typing the information
	name_field = Entry(root)
	ContactNum_field = Entry(root)
	edu_field = Entry(root)
	Institute_field = Entry(root)
	email_id_field = Entry(root)
	address_field = Entry(root)
	Gender_field = Entry(root)
	password_field = Entry(root, show="*")
	Re_enter_password_field = Entry(root, show="*")
	
	# bind method of widget is used for
	# the binding the function with the events

	# whenever the enter key is pressed
	# then call the focus1 function
	name_field.bind("<Return>", focus1)

	# whenever the enter key is pressed
	# then call the focus2 function
	ContactNum_field.bind("<Return>", focus2)

	# whenever the enter key is pressed
	# then call the focus3 function
	edu_field.bind("<Return>", focus3)

	# whenever the enter key is pressed
	# then call the focus4 function
	Institute_field.bind("<Return>", focus4)

	# whenever the enter key is pressed
	# then call the focus5 function
	email_id_field.bind("<Return>", focus5)

	# whenever the enter key is pressed
	# then call the focus6 function
	address_field.bind("<Return>", focus6)

	# whenever the enter key is pressed
	# then call the focus6 function
	Gender_field.bind("<Return>", focus7)

	password_field.bind("<Return>", focus8)

	Re_enter_password_field.bind("<Return>", focus1)

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	name_field.grid(row=1, column=1, ipadx="50")
	ContactNum_field.grid(row=2, column=1, ipadx="50")
	edu_field.grid(row=3, column=1, ipadx="50")
	Institute_field.grid(row=4, column=1, ipadx="50")
	email_id_field.grid(row=5, column=1, ipadx="50")
	address_field.grid(row=6, column=1, ipadx="50")
	Gender_field.grid(row=7, column=1, ipadx="50")
	password_field.grid(row=8, column=1, ipadx="50")
	Re_enter_password_field.grid(row=9, column=1, ipadx="50")

	# call excel function
	excel()

	# create a Submit Button and place into the root window
	submit = Button(root, text="Submit", fg="Black",
							bg="Red", command=insert)
	submit.grid(row=13, column=1)

	# start the GUI
	root.mainloop()
